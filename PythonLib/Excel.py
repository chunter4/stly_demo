import openpyxl as xl


def compare_excel_files(filename1, filename2, startrow, cols2compare):
    diff_found = 0
    
    wb1 = xl.load_workbook(filename1, enumerate)
    ws1 = wb1.worksheets[0]
    ws1_maxrows = ws1.max_row
    ws1_maxcols = ws1.max_column

    wb2 = xl.load_workbook(filename2, enumerate)
    ws2 = wb2.worksheets[0]
    ws2_maxrows = ws2.max_row
    
    # Make sure the row counts match
    if ws1_maxrows != ws2_maxrows:
        return False
    
    for row in range(int(startrow) - 1, ws1_maxrows):
        for col in range(0, ws1_maxcols):
            if str(col) in cols2compare:
                if ws1.cell(row=int(row), column=int(col)).value != ws2.cell(row=int(row), column=int(col)).value:
                    val1 = ws1.cell(row=int(row), column=int(col)).value
                    val2 = ws2.cell(row=int(row), column=int(col)).value
                    
                    print("Data difference found in row " + str(row) + ", column " + str(col) + ". Spreadsheet1 value: "
                          + str(val1) + " != Spreadsheet2 value: " + str(val2))
                    diff_found = 1
    
    if diff_found == 1:
        print("Data differences found.")
        return False
    else:
        print("No data differences found.")
        return True
